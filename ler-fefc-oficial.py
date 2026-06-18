import sys, glob
try:
    from openpyxl import load_workbook
except Exception as e:
    print("FALTA_OPENPYXL:", e); sys.exit(0)

caminhos = glob.glob("**/tse-calculo-de-distribuicao-do-FEFC-2026.xlsx", recursive=True)
if not caminhos:
    print("Arquivo nao encontrado."); sys.exit(0)
arq = caminhos[0]
print("Arquivo:", arq)
wb = load_workbook(arq, data_only=True, read_only=True)
print("Abas:", wb.sheetnames)
for nome in wb.sheetnames:
    ws = wb[nome]
    print("\n===== ABA:", nome, "=====")
    n = 0
    for linha in ws.iter_rows(values_only=True):
        celulas = ["" if v is None else str(v) for v in linha]
        if any(c.strip() for c in celulas):
            print(" | ".join(celulas))
            n += 1
        if n >= 60:
            print("... (cortado em 60 linhas)"); break
